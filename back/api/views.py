from django.shortcuts import render
from rest_framework import generics
from django.contrib.auth.models import User
from django.http import HttpResponse
from .models import *
from .serializer import *
import openpyxl
from rest_framework.response import Response
from rest_framework.decorators import api_view, parser_classes
from rest_framework.generics import ListCreateAPIView, RetrieveUpdateDestroyAPIView
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework.views import APIView
from openpyxl import load_workbook
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.utils import get_column_letter

#Login
class SignUpView(generics.CreateAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializer

#Sensores
#Get e Post
class SensorView(ListCreateAPIView): 
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id','sensor','status']
    search_fields = ['id','sensor','status']


#Put e Del
class SensoresDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer
    permission_classes = [IsAuthenticated]

@api_view(['GET'])
def get_sensor(request):
    return Response(Sensor.TIPO_SENSOR)

@api_view(['GET'])
def get_unidade_medida(request):
    return Response(Sensor.UNIDADE_MEDIDA)

@api_view(['GET'])
def get_status(request):
    return Response(Sensor.STATUS)


#Exportar 
def exportar_sensor_excel(request):
     # Criar planilha
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "sensores"

    # Cabeçalhos
    sheet.append(["sensor", "mac_address", "unidade_med", "latitude","longitude","status"])

    # Dados
    for sensor in Sensor.objects.all():
        sheet.append([
            sensor.sensor,
            sensor.mac_address,
            sensor.unidade_med,
            sensor.latitude,
            sensor.longitude,
            sensor.status,
        ])

    for col in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 7
    # Resposta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=sensores.xlsx'
    workbook.save(response)

    return response


#Ambientes
#Get e Post
class AmbienteView(ListCreateAPIView): 
    queryset = Ambiente.objects.all()
    serializer_class = AmbienteSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['sig']
    search_fields = ['sig']


#Put e Del
class AmbientesDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Ambiente.objects.all()
    serializer_class = AmbienteSerializer
    permission_classes = [IsAuthenticated]

#Exportar
def exportar_ambiente_excel(request):
     # Criar planilha
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "ambientes"

    # Cabeçalhos
    sheet.append(["sig", "descricao", "ni", "responsavel"])

    # Dados
    for ambiente in Ambiente.objects.all():
        sheet.append([
            ambiente.sig,
            ambiente.descricao,
            ambiente.ni,
            ambiente.responsavel,
        ])

    for col in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
    # Resposta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=ambientes.xlsx'
    workbook.save(response)

    return response

#Histórico
#Get e Post
class HistoricoView(ListCreateAPIView): 
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['id','timestamp']
    search_fields = ['id','timestamp']
    def get_queryset(self):
        queryset = super().get_queryset()
        sensor_id = self.request.query_params.get('sensor')
        hora_str = self.request.query_params.get('hora') 
        data_str = self.request.query_params.get('data')

        if sensor_id:
            queryset = queryset.filter(sensor__id=sensor_id)

        if data_str:
            queryset = queryset.filter(timestamp__date=data_str)

        if hora_str:
            try:
                hora_int = int(hora_str)
                queryset = queryset.filter(timestamp__hour=hora_int)            
            except ValueError:
                pass  

        return queryset

#Put e Del
class HistoricosDetailView(RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    permission_classes = [IsAuthenticated]

#Exportar
def exportar_historico_excel(request):
     # Criar planilha
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "historico"

    # Cabeçalhos
    sheet.append(["sensor", "ambiente", "valor", "timestamp"])

    # Dados
    for historico in Historico.objects.all():
        sheet.append([
            str(historico.sensor),
            str(historico.ambiente),
            historico.valor,
            historico.timestamp.strftime('%d/%m/%Y %H:%M:%S'),
        ])

    for col in sheet.columns:
        max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
        sheet.column_dimensions[get_column_letter(col[0].column)].width = max_length + 5
    # Resposta
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename=historico.xlsx'
    workbook.save(response)

    return response    


#Upload Ambientes
class UploadXLSXViewAmbiente(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  # primeira aba
        
        

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            sig = str(row[0]).strip() if row[0] is not None else None
            descricao = str(row[1]).strip() if row[1] is not None else None
            ni = str(row[2]).strip() if row[2] is not None else None
            responsavel = str(row[3]).strip() if row[2] is not None else None
            
            if not ni:
                    print(f"[Linha {i+2}] Erro: ni vazio. Dados: {row}")
                    continue  # pula a linha
                
            Ambiente.objects.create(
                sig=row[0],
                descricao=row[1],
                ni=row[2],
                responsavel=row[3],
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})
    

#Upload Sensores
class UploadXLSXViewSensor(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  # primeira aba
        
        

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            sensor = str(row[0]).strip() if row[0] is not None else None
            mac_address = str(row[1]).strip() if row[1] is not None else None
            unidade_med = str(row[2]).strip() if row[2] is not None else None
            latitude = str(row[3]).strip() if row[2] is not None else None
            longitude = str(row[4]).strip() if row[2] is not None else None
            status = str(row[5]).strip() if row[2] is not None else None
            
            if not mac_address:
                    print(f"[Linha {i+2}] Erro: mac_address vazio. Dados: {row}")
                    continue  # pula a linha
                
            Sensor.objects.create(
                sensor=row[0],
                mac_address=row[1],
                unidade_med=row[2],
                latitude=row[3],
                longitude=row[4],
                status=row[5],
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})
    

#Upload Histórico
class UploadXLSXViewHistorico(APIView):
    parser_classes = [MultiPartParser]

    def post(self, request, *args, **kwargs):
        file_obj = request.FILES.get('file')

        if not file_obj:
            return Response({'erro': 'Arquivo não enviado'}, status=400)

        wb = load_workbook(filename=file_obj)
        ws = wb.active  # primeira aba
        
        

        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):  # pula o cabeçalho
            sensor = row[0]
            ambiente = row[1]
            valor = row[2]
            timestamp = row[3]

            try:
                sensor = Sensor.objects.get(id=sensor)
            except Sensor.DoesNotExist:
                print(f"[Linha {i+2}] Sensor com ID {sensor} não encontrado.")
                continue

            try:
                ambiente = Ambiente.objects.get(id=ambiente)
            except Ambiente.DoesNotExist:
                print(f"[Linha {i+2}] Ambiente com ID {ambiente} não encontrado.")
                continue
                
            Historico.objects.create(
                sensor=sensor,
                ambiente=ambiente,
                valor=valor,
                timestamp=timestamp,
            )

        return Response({'mensagem': 'Dados importados com sucesso!'})

